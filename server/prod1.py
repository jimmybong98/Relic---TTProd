# app.py
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
from threading import Lock
import os
import re
import time

app = Flask(__name__)

# ========= CONFIG =========
# Caminhos originais (rede)
_NET_PREP = r"\\192.168.0.82\00. SGI - Sistema Integrado\12. Qualidade\09. Formulários\For - 007 - Registro de amostragem e For - 008 - Liberação de Maquina 4.xlsx"
_NET_OPER = r"\\192.168.0.82\00. SGI - Sistema Integrado\12. Qualidade\09. Formulários\For - 09 a 14 - Verificação durante o Processo 2.xlsx"

ABA_PREPARADOR = "CADASTRO"
ABA_OPERADOR   = "CADASTRO"

# Onde começam os quartetos de medidas (G/H/I/J -> G é 0-based 6)
COL_MEDIDAS_INICIO = 6
# Chave combinada (A=0 p/ Preparador; C=2 p/ Operador)
COL_CHAVE_COMBINADA_PREPARADOR = 0
COL_CHAVE_COMBINADA_OPERADOR   = 2

# ========= Paths (preferir arquivo local se existir) =========
BASE_DIR = Path(__file__).resolve().parent

def _prefer_local(network_path: str) -> str:
    local = BASE_DIR / Path(network_path).name
    return str(local) if local.exists() else network_path

PLANILHA_PREPARADOR_PATH = _prefer_local(_NET_PREP)
PLANILHA_OPERADOR_PATH   = _prefer_local(_NET_OPER)

# ========= Utils =========
def _norm(text):
    return (str(text or "")).strip()

def _only_digits(text):
    s = re.sub(r"\D+", "", str(text or ""))
    return str(int(s)) if s else ""

def _key_for(part: str, op: str) -> str:
    return f"{_only_digits(part)}*{_only_digits(op)}"

def _to_float(s):
    try:
        return float(str(s).replace(",", ".").strip())
    except Exception:
        return None

def _normalize_text(s: str) -> str:
    t = _norm(s).lower()
    rep = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c"
    }
    for a, b in rep.items():
        t = t.replace(a, b)
    return t

def _has_min_token(s: str) -> bool:
    t = _normalize_text(s)
    return "minimo" in t or re.search(r"\bmin\b", t) is not None

def _has_max_token(s: str) -> bool:
    t = _normalize_text(s)
    return "maximo" in t or re.search(r"\bmax\b", t) is not None

def _is_rugosidade_text(s: str) -> bool:
    t = _normalize_text(s)
    # "rug" + ("ra" ou "rz")
    return ("rug" in t) and ("ra" in t or "rz" in t)

def _first_number(s: str):
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(s))
    return _to_float(m.group(0)) if m else None

def _parse_range_any(texto: str):
    """
    Extrai (min, max, unidade) com tolerância:
      - Faixas: '27,50-28,10', '27.50 ~ 28.10', '27,5 a 28,1', etc.
      - Único valor -> (v, v, uni).
      - Se houver token 'mínimo' (ou 'máximo'), respeita só um lado.
    Unidade: melhor-possível (sufixo final).
    """
    if not texto:
        return (None, None, None)
    s = str(texto)

    # Faixa
    m = re.search(
        r"(-?\d+(?:[.,]\d+)?)\s*(?:-|–|~|a|ate|até|to)\s*(-?\d+(?:[.,]\d+)?)\s*([^\d\s]+.*)?$",
        s, re.IGNORECASE
    )
    if m:
        v1 = _to_float(m.group(1))
        v2 = _to_float(m.group(2))
        uni = (m.group(3) or "").strip() or None
        if v1 is not None and v2 is not None and v1 > v2:
            v1, v2 = v2, v1
        return (v1, v2, uni)

    # Único valor com possíveis tokens
    v = _first_number(s)
    uni_m = re.search(r"[a-zA-Zµ°]+[a-zA-Z0-9/%²³]*$", _norm(s))
    uni = (uni_m.group(0) if uni_m else "").strip() or None

    if v is None:
        return (None, None, None)

    has_min = _has_min_token(s)
    has_max = _has_max_token(s)
    if has_min and not has_max:
        return (v, None, uni)   # mínimo somente
    if has_max and not has_min:
        return (None, v, uni)   # máximo somente
    return (v, v, uni)          # valor exato

def _cell_text(row_vals, col_idx):
    return (
        str(row_vals[col_idx])
        if col_idx < len(row_vals) and row_vals[col_idx] is not None
        else ""
    ).strip()

# ---- Conversão de rótulo de coluna Excel -> índice 0-based ----
def _col_to_idx(label: str) -> int:
    lab = label.strip().upper()
    n = 0
    for ch in lab:
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
        else:
            break
    return n - 1  # 0-based

# Primeiro bloco de tolerâncias começa em AE (0-based 30):
TOL_COL_INICIO = _col_to_idx("AE")  # AE=31 (1-based) => 30 (0-based)

# ========= Cache/Index por aba =========
class SheetIndex:
    """
    Indexa a planilha: key 'part*op' normalizada -> linha (ou lista de linhas).
    Recarrega automaticamente se o arquivo for alterado (mtime).
    """
    def __init__(self, path: str, sheet: str, key_col: int, multi: bool):
        self.path = path
        self.sheet = sheet
        self.key_col = key_col
        self.multi = multi
        self.mtime = None
        self.index = {}
        self.lock = Lock()

    def ensure(self):
        try:
            mtime = os.path.getmtime(self.path)
        except Exception:
            mtime = None

        with self.lock:
            if self.index and self.mtime == mtime:
                return

            t0 = time.time()
            wb = load_workbook(self.path, data_only=True, read_only=True)
            try:
                if self.sheet not in wb.sheetnames:
                    raise RuntimeError(f"Aba '{self.sheet}' não encontrada em {self.path}")
                ws = wb[self.sheet]

                new_index = {}
                for row in ws.iter_rows(values_only=True):
                    cel = row[self.key_col] if self.key_col < len(row) else ""
                    s = _norm(cel)
                    if "*" not in s:
                        continue
                    left, right = s.split("*", 1)
                    key = _key_for(left, right)
                    if not key:
                        continue
                    rv = list(row)
                    if self.multi:
                        new_index.setdefault(key, []).append(rv)
                    else:
                        if key not in new_index:
                            new_index[key] = rv
                self.index = new_index
                self.mtime = mtime
            finally:
                wb.close()

            dt = (time.time() - t0) * 1000
            print(f"[CACHE] Indexado '{self.sheet}' de {Path(self.path).name} em {dt:.0f} ms "
                  f"(chaves: {len(self.index)})", flush=True)

    def get_one(self, key: str):
        self.ensure()
        return self.index.get(key)

    def get_many(self, key: str):
        self.ensure()
        return self.index.get(key, [])

IDX_PREP = SheetIndex(PLANILHA_PREPARADOR_PATH, ABA_PREPARADOR, COL_CHAVE_COMBINADA_PREPARADOR, multi=False)
IDX_OP   = SheetIndex(PLANILHA_OPERADOR_PATH,   ABA_OPERADOR,   COL_CHAVE_COMBINADA_OPERADOR,   multi=True)

# ========= Extração de medidas =========
def _extrair_medidas_pares(row_vals):
    """Para PREPARADOR: pares (etiqueta, faixa)."""
    medidas = []
    col = COL_MEDIDAS_INICIO
    while True:
        etiqueta = _cell_text(row_vals, col)
        faixa    = _cell_text(row_vals, col + 1)
        if not (etiqueta or faixa):
            break

        mn, mx, uni = _parse_range_any(faixa)
        if mn is None and mx is None:
            mn, mx, uni2 = _parse_range_any(etiqueta)
            if uni is None:
                uni = uni2

        # regra rugosidade: valor único vira 0..valor
        if _is_rugosidade_text(etiqueta):
            if mn is not None and mx is not None and mn == mx:
                mx = mn
                mn = 0.0
            elif mn is None and mx is not None:
                mn = 0.0

        medidas.append({
            "titulo": etiqueta or "",
            "faixaTexto": faixa or "",
            "min": mn,
            "max": mx,
            "unidade": uni,
        })
        col += 2
    return medidas

def _extrair_medidas_quartetos(row_vals):
    """
    Para OPERADOR: quartetos (tipo, faixa, periodicidade, instrumento)
    + tolerâncias em blocos de 4 colunas a partir de AE (AE–AH p/ 1º quarteto,
      AI–AL p/ 2º, AM–AP p/ 3º, ...).
    """
    medidas = []
    col = COL_MEDIDAS_INICIO
    idx_quarteto = 0  # 0 para G/H/I/J, 1 para K/L/M/N, 2 para O/P/Q/R, ...

    while True:
        tipo        = _cell_text(row_vals, col)
        faixa       = _cell_text(row_vals, col + 1)
        periodic    = _cell_text(row_vals, col + 2)
        instrumento = _cell_text(row_vals, col + 3)

        if not (tipo or faixa or periodic or instrumento):
            break

        mn, mx, uni = _parse_range_any(faixa)
        if mn is None and mx is None:
            mn, mx, uni2 = _parse_range_any(tipo)
            if uni is None:
                uni = uni2

        # regra rugosidade
        if _is_rugosidade_text(tipo):
            if mn is not None and mx is not None and mn == mx:
                mx = mn
                mn = 0.0
            elif mn is None and mx is not None:
                mn = 0.0

        # ----- Tolerâncias por quarteto (4 colunas cada) -----
        tol_start = TOL_COL_INICIO + idx_quarteto * 4
        tolerancias = []
        for tcol in range(tol_start, tol_start + 4):
            if tcol < len(row_vals):
                d = _to_float(row_vals[tcol])
                if d is not None:
                    tolerancias.append(d)

        medidas.append({
            "titulo": tipo or "",
            "faixaTexto": faixa or "",
            "min": mn,
            "max": mx,
            "unidade": uni,
            "periodicidade": periodic or "",
            "instrumento": instrumento or "",
            "tolerancias": tolerancias,  # pode vir []
        })

        col += 4
        idx_quarteto += 1

    return medidas

# ========= Routes =========
@app.route("/preparador/medidas")
def medidas_preparador():
    part = _norm(request.args.get("partnumber"))
    op   = _norm(request.args.get("operacao"))
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    key = _key_for(part, op)
    print(f"[DEBUG] /preparador/medidas: key={key} (path={PLANILHA_PREPARADOR_PATH})", flush=True)
    try:
        row_vals = IDX_PREP.get_one(key)
        if row_vals is None:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        data = _extrair_medidas_pares(row_vals)
        if not data:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha do PREPARADOR: {e}"}), 500

@app.route("/preparador/resultado", methods=["POST"])
def resultado_preparador():
    data = request.get_json(silent=True) or {}
    print(f"[DEBUG] /preparador/resultado recebido: {data}", flush=True)
    return jsonify({"status": "ok"})

@app.route("/operador/medidas")
def medidas_operador():
    part = _norm(request.args.get("partnumber"))
    op   = _norm(request.args.get("operacao"))
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    key = _key_for(part, op)
    print(f"[DEBUG] /operador/medidas: key={key} (path={PLANILHA_OPERADOR_PATH})", flush=True)
    try:
        linhas = IDX_OP.get_many(key)
        if not linhas:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        medidas = []
        for row_vals in linhas:
            medidas.extend(_extrair_medidas_quartetos(row_vals))
        if not medidas:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        return jsonify(medidas)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha do OPERADOR: {e}"}), 500

if __name__ == "__main__":
    # threaded=True mantém atendendo enquanto indexa em background entre requests
    app.run(host="0.0.0.0", port=5005, debug=True, threaded=True, use_reloader=False)
