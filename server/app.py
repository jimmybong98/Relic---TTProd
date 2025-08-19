# server/app.py
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import re

app = Flask(__name__)

# ========================= CONFIG =========================
PLANILHA_PREPARADOR_PATH = (
    r"\\\\192.168.0.82\\00. SGI - Sistema Integrado\\12. Qualidade\\09. Formulários\\For - 007 - Registro de amostragem e For - 008 - Liberação de Maquina 4.xlsx"
)
PLANILHA_OPERADOR_PATH = (
    r"\\\\192.168.0.82\\00. SGI - Sistema Integrado\\12. Qualidade\\09. Formulários\\For - 09 a 14 - Verificação durante o Processo.xlsx"
)
ABA_PREPARADOR = "CADASTRO"
ABA_OPERADOR = "CADASTRO"

# Índices 0-based
COL_CHAVE_PREPARADOR = 0  # A
COL_CHAVE_OPERADOR = 2    # C
COL_MEDIDAS_INICIO = 6    # G

# ========================= HELPERS ========================
_NON_DIGITS = re.compile(r'\D+')
_RANGE_RE = re.compile(r"""
    ^\s*
    (?P<min>[-+]?\d+(?:[.,]\d+)?)
    \s*[-–]\s*
    (?P<max>[-+]?\d+(?:[.,]\d+)?)
    (?:\s*(?P<uni>mm|cm|in|º|°|[a-zA-Z%]+))?
    \s*$
""", re.X)

def _cell_text(row, idx0: int) -> str:
    if idx0 < 0 or idx0 >= len(row):
        return ""
    v = row[idx0]
    return "" if v is None else str(v).strip()

def _normalize_key_strict(part: str, op: str) -> str:
    # igualdade exata: "000000000373*010"
    return f"{part}*{op}"

def _only_digits(s: str) -> str:
    return _NON_DIGITS.sub('', s or '')

def _norm_int(s: str) -> int:
    ds = _only_digits(s)
    return int(ds) if ds else -1

def _keys_match_fallback(cell_value: str, part: str, op: str) -> bool:
    """
    Fallback tolerante: ignora zeros à esquerda e espaços.
    Casa se dígitos(part) == dígitos(PN da célula) e dígitos(op) == dígitos(OP da célula).
    """
    s = (cell_value or "").strip()
    if not s or "*" not in s:
        return False
    left, right = s.split("*", 1)
    return _norm_int(left) == _norm_int(part) and _norm_int(right) == _norm_int(op)

def _parse_range(texto: str):
    if not texto:
        return (None, None, None)
    m = _RANGE_RE.match(str(texto))
    if not m:
        return (None, None, None)
    def to_float(x):
        return float(str(x).replace(',', '.')) if x else None
    return (to_float(m.group('min')), to_float(m.group('max')), m.group('uni'))

def _extrair_medidas_da_linha(row):
    medidas = []
    col = COL_MEDIDAS_INICIO  # G
    while True:
        etiqueta = _cell_text(row, col)
        especific = _cell_text(row, col + 1)
        if not etiqueta and not especific:
            break
        # Apenas exibição (escopo atual). min/max/unidade opcionais.
        mn, mx, uni = _parse_range(especific)
        medidas.append({
            "titulo": etiqueta or "",
            "faixa": especific or "",
            "min": mn,
            "max": mx,
            "unidade": uni,
        })
        col += 2
    return medidas

def _encontrar_linha(ws, part: str, op: str, col_chave: int):
    """
    Procura primeiro por igualdade EXATA na coluna indicada.
    Se não encontrar, usa fallback tolerante na mesma coluna.
    Retorna (row_values) ou None.
    """
    chave_exata = _normalize_key_strict(part, op)

    # 1) Igualdade exata
    for row in ws.iter_rows(values_only=True):
        a_val = _cell_text(row, col_chave)
        if a_val == chave_exata:
            return row

    # 2) Fallback tolerante (zeros à esquerda, espaços)
    for row in ws.iter_rows(values_only=True):
        a_val = _cell_text(row, col_chave)
        if _keys_match_fallback(a_val, part, op):
            return row

    return None

# ========================= ROUTES =========================
@app.get("/health")
def health():
    ok_prep = Path(PLANILHA_PREPARADOR_PATH).exists()
    ok_oper = Path(PLANILHA_OPERADOR_PATH).exists()
    return jsonify(
        {
            "preparador": {
                "ok": ok_prep,
                "path": PLANILHA_PREPARADOR_PATH,
                "sheet": ABA_PREPARADOR,
            },
            "operador": {
                "ok": ok_oper,
                "path": PLANILHA_OPERADOR_PATH,
                "sheet": ABA_OPERADOR,
            },
        }
    )

@app.get("/medidas")
def get_medidas():
    """
    GET /medidas?partnumber=000000000373&operacao=010
    Retorna: [{"titulo":"DIAMETRO","faixa":"4.05-4.20", "min":4.05, "max":4.20, "unidade":"mm"}, ...]
    (min/max/unidade podem ser None se a faixa não tiver padrão numérico)
    """
    part = (request.args.get("partnumber") or "").strip()
    op = (request.args.get("operacao") or "").strip()
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    try:
        p = Path(PLANILHA_PREPARADOR_PATH)
        if not p.exists():
            return jsonify({"error": f"Planilha não encontrada: {PLANILHA_PREPARADOR_PATH}"}), 500

        wb = load_workbook(PLANILHA_PREPARADOR_PATH, data_only=True, read_only=True)
        if ABA_PREPARADOR not in wb.sheetnames:
            return jsonify({"error": f"Aba '{ABA_PREPARADOR}' não encontrada"}), 500
        ws = wb[ABA_PREPARADOR]

        row = _encontrar_linha(ws, part, op, COL_CHAVE_PREPARADOR)
        if row is None:
            return jsonify([])  # não encontrou a chave

        medidas = _extrair_medidas_da_linha(row)
        return jsonify(medidas)

    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha: {e}"}), 500

@app.post("/medidas/resultado")
def post_resultado():
    """
    Persistência fica para a próxima etapa.
    Apenas confirma recebimento do payload.
    """
    data = request.get_json(silent=True) or {}
    obrig = ["re", "partnumber", "operacao", "itens"]
    faltando = [k for k in obrig if k not in data]
    if faltando:
        return jsonify({"error": f"Campos obrigatórios ausentes: {faltando}"}), 400
    return jsonify({"ok": True, "received_at": datetime.utcnow().isoformat() + "Z"})

# --------------------- Operador ---------------------------

@app.get("/operador/medidas")
def get_medidas_operador():
    part = (request.args.get("partnumber") or "").strip()
    op = (request.args.get("operacao") or "").strip()
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    try:
        p = Path(PLANILHA_OPERADOR_PATH)
        if not p.exists():
            return jsonify({"error": f"Planilha não encontrada: {PLANILHA_OPERADOR_PATH}"}), 500

        wb = load_workbook(PLANILHA_OPERADOR_PATH, data_only=True, read_only=True)
        if ABA_OPERADOR not in wb.sheetnames:
            return jsonify({"error": f"Aba '{ABA_OPERADOR}' não encontrada"}), 500
        ws = wb[ABA_OPERADOR]

        row = _encontrar_linha(ws, part, op, COL_CHAVE_OPERADOR)
        if row is None:
            return jsonify([])

        medidas = _extrair_medidas_da_linha(row)
        return jsonify(medidas)

    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha: {e}"}), 500

# ========================= MAIN ==========================
if __name__ == "__main__":
    # Ex.: http://localhost:5005/medidas?partnumber=000000000373&operacao=010
    app.run(host="0.0.0.0", port=5005, debug=False)
