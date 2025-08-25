# app.py
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from pathlib import Path
from threading import Lock
from typing import Optional, Tuple
import os
import re
import time
import json

# --------- MySQL ----------
import pymysql
from pymysql.cursors import DictCursor

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False  # Flask 2.x
try:  # Flask 3.x
    app.json.ensure_ascii = False
except Exception:
    pass

# ========= CONFIG =========
# Caminhos originais (rede)
_NET_PREP = r"\\192.168.0.82\00. SGI - Sistema Integrado\12. Qualidade\09. Formulários\For - 007 - Registro de amostragem e For - 008 - Liberação de Maquina 4.xlsx"
_NET_OPER = r"\\192.168.0.82\00. SGI - Sistema Integrado\12. Qualidade\09. Formulários\For - 09 a 14 - Verificação durante o Processo 2.xlsx"

ABA_PREPARADOR = "CADASTRO"
ABA_OPERADOR   = "CADASTRO"

# Onde começam as medidas (pares para preparador / quartetos para operador)
COL_MEDIDAS_INICIO = 6  # G em 0-based

# Chave combinada (A=0 p/ Preparador; C=2 p/ Operador)
COL_CHAVE_COMBINADA_PREPARADOR = 0
COL_CHAVE_COMBINADA_OPERADOR   = 2

# ========= Paths (preferir arquivo local se existir) =========
BASE_DIR = Path(__file__).resolve().parent

def _prefer_local(network_path: str) -> str:
    local = BASE_DIR / Path(network_path).name
    return str(local) if local.exists() else network_path

PLANILHA_PREPARADOR_PATH = _prefer_local(_NET_PREP)
PLANILHA_OPERADOR_PATH   = _prefer_local(_NET_OPER)

# ========= DB (MySQL) =========
DB_HOST = os.getenv("DB_HOST", "192.168.0.31")
DB_PORT = int(os.getenv("DB_PORT", "3306"))
DB_USER = os.getenv("DB_USER", "relic")
DB_PASS = os.getenv("DB_PASS", "veALZ2FBnDkG749")
DB_NAME = os.getenv("DB_NAME", "relic_quality")

def _conn_db(dbname: Optional[str] = None):
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASS,
        database=dbname,
        autocommit=False,
        cursorclass=DictCursor,
        charset="utf8mb4",
    )

def _ensure_schema():
    """Garante que o banco e as tabelas principais existam (sem DDL agressivo)."""
    # Cria o database, se não existir
    with _conn_db(None) as c:
        with c.cursor() as cur:
            cur.execute(
                f"CREATE DATABASE IF NOT EXISTS `{DB_NAME}` "
                "CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;"
            )
        c.commit()

    # Cria tabelas mínimas usadas aqui (se já existem, não mexe)
    with _conn_db(DB_NAME) as c:
        with c.cursor() as cur:
            # Tabela-mestra de OS
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS ordem_servico (
                  os VARCHAR(64) NOT NULL,
                  descricao VARCHAR(255) DEFAULT NULL,
                  cliente VARCHAR(255) DEFAULT NULL,
                  created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  atualizado_em TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  PRIMARY KEY (os)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            # Operador (já estava)
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS operador_amostragem (
                    id BIGINT AUTO_INCREMENT PRIMARY KEY,
                    os VARCHAR(64) NOT NULL,
                    partnumber VARCHAR(128) NOT NULL,
                    operacao VARCHAR(64) NOT NULL,
                    re_operador VARCHAR(64) NOT NULL,
                    observacao TEXT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    KEY idx_oa_os (os),
                    KEY idx_oa_part_op (partnumber, operacao),
                    KEY idx_oa_created (created_at),
                    CONSTRAINT fk_oa_os FOREIGN KEY (os) REFERENCES ordem_servico(os) ON UPDATE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS operador_amostragem_item (
                    id BIGINT AUTO_INCREMENT PRIMARY KEY,
                    amostragem_id BIGINT NOT NULL,
                    idx_medida INT NOT NULL,
                    titulo TEXT,
                    instrumento VARCHAR(255),
                    faixa_texto TEXT,
                    minimo DOUBLE NULL,
                    maximo DOUBLE NULL,
                    unidade VARCHAR(64) NULL,
                    periodicidade VARCHAR(128) NULL,
                    tolerancias LONGTEXT
                      CHARACTER SET utf8mb4
                      COLLATE utf8mb4_bin
                      NULL,
                    escolha VARCHAR(128) NOT NULL,
                    status VARCHAR(64) NULL,
                    observacao TEXT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_oa_item (amostragem_id, idx_medida),
                    CONSTRAINT fk_oa_item_oa
                        FOREIGN KEY (amostragem_id)
                        REFERENCES operador_amostragem(id)
                        ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            # Preparador (registro + itens)
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS preparador_registro (
                  id BIGINT AUTO_INCREMENT PRIMARY KEY,
                  os VARCHAR(64) NOT NULL,
                  partnumber VARCHAR(128) NOT NULL,
                  operacao VARCHAR(64) NOT NULL,
                  re_preparador VARCHAR(64) NOT NULL,
                  created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  KEY idx_os (os),
                  KEY idx_part_op (partnumber, operacao)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS preparador_registro_item (
                  id BIGINT AUTO_INCREMENT PRIMARY KEY,
                  registro_id BIGINT NOT NULL,
                  idx_medida INT NOT NULL,
                  titulo TEXT DEFAULT NULL,
                  faixa_texto TEXT DEFAULT NULL,
                  minimo DOUBLE DEFAULT NULL,
                  maximo DOUBLE DEFAULT NULL,
                  unidade VARCHAR(64) DEFAULT NULL,
                  medicao TEXT DEFAULT NULL,
                  status VARCHAR(64) DEFAULT NULL,
                  observacao TEXT DEFAULT NULL,
                  created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  KEY idx_cab (registro_id),
                  KEY idx_idx (idx_medida),
                  CONSTRAINT fk_prep_registro
                    FOREIGN KEY (registro_id)
                    REFERENCES preparador_registro(id)
                    ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            # Preparador (liberação consolidada)
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS preparador_liberacao (
                  id BIGINT AUTO_INCREMENT PRIMARY KEY,
                  os VARCHAR(64) NOT NULL,
                  partnumber VARCHAR(128) NOT NULL,
                  operacao VARCHAR(64) NOT NULL,
                  re_preparador VARCHAR(64) NOT NULL,
                  status_geral VARCHAR(32) DEFAULT NULL,
                  observacao TEXT DEFAULT NULL,
                  created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  KEY idx_pl_os (os),
                  KEY idx_pl_part_op (partnumber, operacao),
                  KEY idx_pl_created (created_at)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS preparador_liberacao_item (
                  id BIGINT AUTO_INCREMENT PRIMARY KEY,
                  liberacao_id BIGINT NOT NULL,
                  idx_medida INT NOT NULL,
                  titulo TEXT DEFAULT NULL,
                  faixa_texto TEXT DEFAULT NULL,
                  minimo DOUBLE DEFAULT NULL,
                  maximo DOUBLE DEFAULT NULL,
                  unidade VARCHAR(64) DEFAULT NULL,
                  medicao DOUBLE DEFAULT NULL,
                  status VARCHAR(64) NOT NULL,
                  periodicidade VARCHAR(128) DEFAULT NULL,
                  instrumento VARCHAR(255) DEFAULT NULL,
                  observacao TEXT DEFAULT NULL,
                  created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  UNIQUE KEY uq_pl_item (liberacao_id, idx_medida),
                  CONSTRAINT fk_pl_item_pl
                    FOREIGN KEY (liberacao_id)
                    REFERENCES preparador_liberacao(id)
                    ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )
        c.commit()

# chama ao subir
_ensure_schema()

# ========= Utils =========
def _norm(text):
    return (str(text or "")).strip()

def _only_digits(text):
    s = re.sub(r"\D+", "", str(text or ""))
    return str(int(s)) if s else ""

def _key_for(part: str, op: str) -> str:
    return f"{_only_digits(part)}*{_only_digits(op)}"

def _to_float(s):
    try:
        return float(str(s).replace(",", ".").strip())
    except Exception:
        return None

def _normalize_text(s: str) -> str:
    t = _norm(s).lower()
    rep = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c"
    }
    for a, b in rep.items():
        t = t.replace(a, b)
    return t

def _has_min_token(s: str) -> bool:
    t = _normalize_text(s)
    return "minimo" in t or re.search(r"\bmin\b", t) is not None

def _has_max_token(s: str) -> bool:
    t = _normalize_text(s)
    return "maximo" in t or re.search(r"\bmax\b", t) is not None

def _is_rugosidade_text(s: str) -> bool:
    t = _normalize_text(s)
    # "rug" + ("ra" ou "rz")
    return ("rug" in t) and ("ra" in t or "rz" in t)

def _first_number(s: str):
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(s))
    return _to_float(m.group(0)) if m else None

def _parse_range_any(texto: str):
    """
    Extrai (min, max, unidade) com tolerância:
      - Faixas: '27,50-28,10', '27.50 ~ 28.10', '27,5 a 28,1', etc.
      - Único valor -> (v, v, uni).
      - Se houver token 'mínimo' (ou 'máximo'), respeita só um lado.
    Unidade: melhor-possível (sufixo final).
    """
    if not texto:
        return (None, None, None)
    s = str(texto)

    # Faixa
    m = re.search(
        r"(-?\d+(?:[.,]\d+)?)\s*(?:-|–|~|a|ate|até|to)\s*(-?\d+(?:[.,]\d+)?)\s*([^\d\s]+.*)?$",
        s, re.IGNORECASE
    )
    if m:
        v1 = _to_float(m.group(1))
        v2 = _to_float(m.group(2))
        uni = (m.group(3) or "").strip() or None
        if v1 is not None and v2 is not None and v1 > v2:
            v1, v2 = v2, v1
        return (v1, v2, uni)

    # Único valor com possíveis tokens
    v = _first_number(s)
    uni_m = re.search(r"[a-zA-Zµ°]+[a-zA-Z0-9/%²³]*$", _norm(s))
    uni = (uni_m.group(0) if uni_m else "").strip() or None

    if v is None:
        return (None, None, None)

    has_min = _has_min_token(s)
    has_max = _has_max_token(s)
    if has_min and not has_max:
        return (v, None, uni)   # mínimo somente
    if has_max and not has_min:
        return (None, v, uni)   # máximo somente
    return (v, v, uni)          # valor exato

def _cell_text(row_vals, col_idx):
    return (
        str(row_vals[col_idx])
        if col_idx < len(row_vals) and row_vals[col_idx] is not None
        else ""
    ).strip()

# ---- Conversão de rótulo de coluna Excel -> índice 0-based ----
def _col_to_idx(label: str) -> int:
    lab = label.strip().upper()
    n = 0
    for ch in lab:
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
        else:
            break
    return n - 1  # 0-based

# Primeiro bloco de tolerâncias começa em AE (0-based 30):
TOL_COL_INICIO = _col_to_idx("AE")  # AE=31 (1-based) => 30 (0-based)

# ========= Cache/Index por aba =========
class SheetIndex:
    """
    Indexa a planilha: key 'part*op' normalizada -> linha (ou lista de linhas).
    Recarrega automaticamente se o arquivo for alterado (mtime).
    """
    def __init__(self, path: str, sheet: str, key_col: int, multi: bool):
        self.path = path
        self.sheet = sheet
        self.key_col = key_col
        self.multi = multi
        self.mtime = None
        self.index = {}
        self.lock = Lock()

    def ensure(self):
        try:
            mtime = os.path.getmtime(self.path)
        except Exception:
            mtime = None

        with self.lock:
            if self.index and self.mtime == mtime:
                return

            t0 = time.time()
            wb = load_workbook(self.path, data_only=True, read_only=True)
            try:
                if self.sheet not in wb.sheetnames:
                    raise RuntimeError(f"Aba '{self.sheet}' não encontrada em {self.path}")
                ws = wb[self.sheet]

                new_index = {}
                for row in ws.iter_rows(values_only=True):
                    cel = row[self.key_col] if self.key_col < len(row) else ""
                    s = _norm(cel)
                    if "*" not in s:
                        continue
                    left, right = s.split("*", 1)
                    key = _key_for(left, right)
                    if not key:
                        continue
                    rv = list(row)
                    if self.multi:
                        new_index.setdefault(key, []).append(rv)
                    else:
                        if key not in new_index:
                            new_index[key] = rv
                self.index = new_index
                self.mtime = mtime
            finally:
                wb.close()

            dt = (time.time() - t0) * 1000
            print(f"[CACHE] Indexado '{self.sheet}' de {Path(self.path).name} em {dt:.0f} ms "
                  f"(chaves: {len(self.index)})", flush=True)

    def get_one(self, key: str):
        self.ensure()
        return self.index.get(key)

    def get_many(self, key: str):
        self.ensure()
        return self.index.get(key, [])

IDX_PREP = SheetIndex(PLANILHA_PREPARADOR_PATH, ABA_PREPARADOR, COL_CHAVE_COMBINADA_PREPARADOR, multi=False)
IDX_OP   = SheetIndex(PLANILHA_OPERADOR_PATH,   ABA_OPERADOR,   COL_CHAVE_COMBINADA_OPERADOR,   multi=True)

# ========= Extração de medidas =========
def _extrair_medidas_pares(row_vals):
    """Para PREPARADOR: pares (etiqueta, faixa)."""
    medidas = []
    col = COL_MEDIDAS_INICIO
    while True:
        etiqueta = _cell_text(row_vals, col)
        faixa    = _cell_text(row_vals, col + 1)
        if not (etiqueta or faixa):
            break

        mn, mx, uni = _parse_range_any(faixa)
        if mn is None and mx is None:
            mn, mx, uni2 = _parse_range_any(etiqueta)
            if uni is None:
                uni = uni2

        # regra rugosidade: valor único vira 0..valor
        if _is_rugosidade_text(etiqueta):
            if mn is not None and mx is not None and mn == mx:
                mx = mn
                mn = 0.0
            elif mn is None and mx is not None:
                mn = 0.0

        medidas.append({
            "titulo": etiqueta or "",
            "faixaTexto": faixa or "",
            "min": mn,
            "max": mx,
            "unidade": uni,
        })
        col += 2
    return medidas

def _extrair_medidas_quartetos(row_vals):
    """
    Para OPERADOR: quartetos (tipo, faixa, periodicidade, instrumento)
    + tolerâncias em blocos de 4 colunas a partir de AE (AE–AH p/ 1º quarteto,
      AI–AL p/ 2º, AM–AP p/ 3º, ...).
    """
    medidas = []
    col = COL_MEDIDAS_INICIO
    idx_quarteto = 0  # 0 para G/H/I/J, 1 para K/L/M/N, 2 para O/P/Q/R, ...

    while True:
        tipo        = _cell_text(row_vals, col)
        faixa       = _cell_text(row_vals, col + 1)
        periodic    = _cell_text(row_vals, col + 2)
        instrumento = _cell_text(row_vals, col + 3)

        if not (tipo or faixa or periodic or instrumento):
            break

        mn, mx, uni = _parse_range_any(faixa)
        if mn is None and mx is None:
            mn, mx, uni2 = _parse_range_any(tipo)
            if uni is None:
                uni = uni2

        # regra rugosidade
        if _is_rugosidade_text(tipo):
            if mn is not None and mx is not None and mn == mx:
                mx = mn
                mn = 0.0
            elif mn is None and mx is not None:
                mn = 0.0

        # ----- Tolerâncias por quarteto (4 colunas cada) -----
        tol_start = TOL_COL_INICIO + idx_quarteto * 4
        tolerancias = []
        for tcol in range(tol_start, tol_start + 4):
            if tcol < len(row_vals):
                d = _to_float(row_vals[tcol])
                if d is not None:
                    tolerancias.append(d)

        medidas.append({
            "titulo": tipo or "",
            "faixaTexto": faixa or "",
            "min": mn,
            "max": mx,
            "unidade": uni,
            "periodicidade": periodic or "",
            "instrumento": instrumento or "",
            "tolerancias": tolerancias,  # pode vir []
        })

        col += 4
        idx_quarteto += 1

    return medidas

# ========= HELPERS DE NEGÓCIO =========
def _maquina_liberada(conn, os_num: str, part: str, op: str) -> Tuple[bool, str, str]:
    """
    Retorna (liberada, fonte, detalhe).
    fonte: 'preparador_liberacao' | 'preparador_registro' | ''
    """
    os_num = _norm(os_num)
    part   = _norm(part)
    op     = _norm(op)
    if not (os_num and part and op):
        return (False, "", "Parâmetros insuficientes para validação.")

    with conn.cursor() as cur:
        # 1) Se existir liberação com status final, já libera
        cur.execute(
            """
            SELECT status_geral
            FROM preparador_liberacao
            WHERE os=%s AND partnumber=%s AND operacao=%s
            ORDER BY id DESC LIMIT 1
            """,
            (os_num, part, op)
        )
        row = cur.fetchone()
        if row:
            st = (row.get("status_geral") or "").strip().lower()
            if st in ("liberada", "liberado", "ok", "aprovada", "aprovado"):
                return (True, "preparador_liberacao", f"status_geral={st}")
            # se há registro mas não liberada, informa
            return (False, "preparador_liberacao", f"status_geral={st or 'indefinido'}")

        # 2) Caso não tenha liberação, checa o último registro do preparador:
        cur.execute(
            """
            SELECT id
            FROM preparador_registro
            WHERE os=%s AND partnumber=%s AND operacao=%s
            ORDER BY created_at DESC, id DESC LIMIT 1
            """,
            (os_num, part, op)
        )
        reg = cur.fetchone()
        if not reg:
            return (False, "", "Sem registro do preparador para esta OS/peça/operação.")

        reg_id = reg["id"]
        cur.execute(
            """
            SELECT
              SUM(CASE WHEN LOWER(COALESCE(status,''))='ok' THEN 1 ELSE 0 END) AS ok_cnt,
              COUNT(*) AS total
            FROM preparador_registro_item
            WHERE registro_id=%s
            """,
            (reg_id,)
        )
        stats = cur.fetchone() or {}
        ok_cnt = int(stats.get("ok_cnt") or 0)
        total  = int(stats.get("total") or 0)
        if total > 0 and ok_cnt == total:
            return (True, "preparador_registro", f"registro_id={reg_id}; {ok_cnt}/{total} OK")
        else:
            return (False, "preparador_registro", f"registro_id={reg_id}; {ok_cnt}/{total} OK")

# ========= Rotas de Leitura =========
@app.route("/preparador/medidas")
def medidas_preparador():
    part = _norm(request.args.get("partnumber"))
    op   = _norm(request.args.get("operacao"))
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    key = _key_for(part, op)
    print(f"[DEBUG] /preparador/medidas: key={key} (path={PLANILHA_PREPARADOR_PATH})", flush=True)
    try:
        row_vals = IDX_PREP.get_one(key)
        if row_vals is None:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        data = _extrair_medidas_pares(row_vals)
        if not data:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha do PREPARADOR: {e}"}), 500

@app.route("/operador/medidas")
def medidas_operador():
    part = _norm(request.args.get("partnumber"))
    op   = _norm(request.args.get("operacao"))
    if not part or not op:
        return jsonify({"error": "Parâmetros 'partnumber' e 'operacao' são obrigatórios"}), 400

    key = _key_for(part, op)
    print(f"[DEBUG] /operador/medidas: key={key} (path={PLANILHA_OPERADOR_PATH})", flush=True)
    try:
        linhas = IDX_OP.get_many(key)
        if not linhas:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        medidas = []
        for row_vals in linhas:
            medidas.extend(_extrair_medidas_quartetos(row_vals))
        if not medidas:
            return jsonify({"error": "Nenhuma medida encontrada para os parâmetros informados"}), 404
        return jsonify(medidas)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha do OPERADOR: {e}"}), 500

# ========= Registro (MySQL): PREPARADOR =========
@app.route("/preparador/resultado", methods=["POST"])
def resultado_preparador():
    """
    Recebe o registro do PREPARADOR e grava no MySQL.
    Payload:
    {
      "os": "...", "re": "...", "partnumber": "...", "operacao": "...",
      "itens": [
        {
          "indice": 0,
          "titulo": "...",
          "faixaTexto": "...",
          "min": 1.23, "max": 4.56, "unidade": "mm",
          "medicao": "1.30",
          "status": "ok|reprovada_acima|reprovada_abaixo|alerta|pendente",
          "observacao": ""
        }, ...
      ]
    }
    """
    payload = request.get_json(silent=True) or {}
    print(f"[DEBUG] /preparador/resultado recebido: {payload}", flush=True)

    os_num = _norm(payload.get("os"))
    re_prep = _norm(payload.get("re"))
    part = _norm(payload.get("partnumber"))
    op = _norm(payload.get("operacao"))
    itens = payload.get("itens", [])

    if not os_num or not re_prep or not part or not op:
        return jsonify({"error": "Campos 'os', 're', 'partnumber' e 'operacao' são obrigatórios"}), 400
    if not isinstance(itens, list) or len(itens) == 0:
        return jsonify({"error": "Lista 'itens' é obrigatória e não pode ser vazia"}), 400

    try:
        with _conn_db(DB_NAME) as c:
            with c.cursor() as cur:
                # garante OS na mestre (por causa de FKs futuras)
                cur.execute("INSERT IGNORE INTO ordem_servico (os) VALUES (%s)", (os_num,))

                # cabeçalho
                cur.execute(
                    """
                    INSERT INTO preparador_registro (os, partnumber, operacao, re_preparador)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (os_num, part, op, re_prep)
                )
                registro_id = cur.lastrowid

                # itens
                all_status = []
                for it in itens:
                    idx = int(it.get("indice", 0))
                    titulo = _norm(it.get("titulo"))
                    faixa_texto = _norm(it.get("faixaTexto"))
                    minimo = it.get("min")
                    maximo = it.get("max")
                    unidade = _norm(it.get("unidade"))
                    medicao = _norm(it.get("medicao"))
                    status = _norm(it.get("status")).lower()
                    observacao = _norm(it.get("observacao"))

                    cur.execute(
                        """
                        INSERT INTO preparador_registro_item
                          (registro_id, idx_medida, titulo, faixa_texto, minimo, maximo, unidade,
                           medicao, status, observacao)
                        VALUES
                          (%s, %s, %s, %s, %s, %s, %s,
                           %s, %s, %s)
                        """,
                        (
                            registro_id, idx, titulo, faixa_texto, minimo, maximo, unidade,
                            medicao, status, observacao
                        )
                    )
                    all_status.append(status)

                # Consolida liberação
                has_reprov = any(s.startswith("reprovada") for s in all_status)
                all_ok = len(all_status) > 0 and all(s == "ok" for s in all_status)
                status_geral = "liberada" if all_ok else ("reprovada" if has_reprov else "pendente")

                # upsert simples em preparador_liberacao (não cria itens aqui)
                cur.execute(
                    """
                    SELECT id FROM preparador_liberacao
                    WHERE os=%s AND partnumber=%s AND operacao=%s
                    ORDER BY id DESC LIMIT 1
                    """,
                    (os_num, part, op)
                )
                row = cur.fetchone()
                if row:
                    cur.execute(
                        "UPDATE preparador_liberacao SET re_preparador=%s, status_geral=%s WHERE id=%s",
                        (re_prep, status_geral, row["id"])
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO preparador_liberacao
                          (os, partnumber, operacao, re_preparador, status_geral)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (os_num, part, op, re_prep, status_geral)
                    )

            c.commit()

        return jsonify({"status": "ok", "registro_id": registro_id, "status_geral": status_geral})

    except Exception as e:
        return jsonify({"error": f"Falha ao inserir registro do preparador: {e}"}), 500

# ========= CHECAGEM (para UI) =========
@app.route("/operador/pode")
def operador_pode():
    """Consulta rápida: pode o operador amostrar? (máquina liberada?)"""
    os_num = _norm(request.args.get("os"))
    part = _norm(request.args.get("partnumber"))
    op = _norm(request.args.get("operacao"))
    if not os_num or not part or not op:
        return jsonify({"error": "Parâmetros 'os', 'partnumber' e 'operacao' são obrigatórios"}), 400

    try:
        with _conn_db(DB_NAME) as c:
            ok, fonte, detalhe = _maquina_liberada(c, os_num, part, op)
        return jsonify({"os": os_num, "partnumber": part, "operacao": op, "liberada": ok, "fonte": fonte, "detalhe": detalhe})
    except Exception as e:
        return jsonify({"error": f"Falha ao consultar liberação: {e}"}), 500

# ========= Registro (MySQL): OPERADOR =========
@app.route("/operador/registrar", methods=["POST"])
def operador_registrar():
    """
    Recebe a amostragem do Operador para gravar no MySQL.
    BLOQUEIA se máquina não estiver liberada pelo preparador.

    Payload:
    {
      "os": "...", "re": "...", "partnumber": "...", "operacao": "...",
      "itens": [
        {
          "indice": 0, "titulo": "...", "instrumento": "...",
          "faixaTexto": "...", "min": 1.23, "max": 4.56, "unidade": "mm",
          "periodicidade": "5 peças", "tolerancias": [..],
          "escolha": "OK", "status": "ok|reprovada_acima|reprovada_abaixo|alerta",
          "observacao": "..."
        }, ...
      ]
    }
    """
    payload = request.get_json(silent=True) or {}

    os_num = _norm(payload.get("os"))
    re_op = _norm(payload.get("re"))
    part = _norm(payload.get("partnumber"))
    op = _norm(payload.get("operacao"))
    itens = payload.get("itens", [])

    # validações mínimas
    if not os_num or not re_op or not part or not op:
        return jsonify({"error": "Campos 'os', 're', 'partnumber' e 'operacao' são obrigatórios"}), 400
    if not isinstance(itens, list) or len(itens) == 0:
        return jsonify({"error": "Lista 'itens' é obrigatória e não pode ser vazia"}), 400

    try:
        with _conn_db(DB_NAME) as c:
            # BLOQUEIO: exige liberação
            ok, fonte, detalhe = _maquina_liberada(c, os_num, part, op)
            if not ok:
                msg = _mensagem_bloqueio(os_num, part, op, fonte, detalhe)
                return jsonify({
                    "code": "liberacao_pendente",
                    "error": msg,                 # amigável (mantive a chave 'error' p/ compatibilidade)
                    "fonte": fonte,
                    "detalhe": detalhe
                }), 409

            with c.cursor() as cur:
                # garante OS na mestre (FK)
                cur.execute("INSERT IGNORE INTO ordem_servico (os) VALUES (%s)", (os_num,))

                # cabeçalho
                cur.execute(
                    """
                    INSERT INTO operador_amostragem (os, partnumber, operacao, re_operador)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (os_num, part, op, re_op)
                )
                amostragem_id = cur.lastrowid

                # itens
                for it in itens:
                    idx = int(it.get("indice", 0))
                    titulo = _norm(it.get("titulo"))
                    instrumento = _norm(it.get("instrumento"))
                    faixa_texto = _norm(it.get("faixaTexto"))
                    minimo = it.get("min")
                    maximo = it.get("max")
                    unidade = _norm(it.get("unidade"))
                    periodicidade = _norm(it.get("periodicidade"))
                    tolerancias = it.get("tolerancias", [])
                    tol_txt = None
                    if isinstance(tolerancias, (list, tuple)):
                        try:
                            tol_txt = json.dumps(tolerancias, ensure_ascii=False)
                        except Exception:
                            tol_txt = None
                    escolha = _norm(it.get("escolha"))
                    status = _norm(it.get("status"))
                    observacao = _norm(it.get("observacao"))

                    cur.execute(
                        """
                        INSERT INTO operador_amostragem_item
                          (amostragem_id, idx_medida, titulo, instrumento, faixa_texto,
                           minimo, maximo, unidade, periodicidade, tolerancias,
                           escolha, status, observacao)
                        VALUES
                          (%s, %s, %s, %s, %s,
                           %s, %s, %s, %s, %s,
                           %s, %s, %s)
                        """,
                        (
                            amostragem_id, idx, titulo, instrumento, faixa_texto,
                            minimo, maximo, unidade, periodicidade, tol_txt,
                            escolha, status, observacao
                        )
                    )
            c.commit()

        return jsonify({"status": "ok", "amostragem_id": amostragem_id, "itens": len(itens)})

    except Exception as e:
        return jsonify({"error": f"Falha ao inserir amostragem: {e}"}), 500

# (Opcional) listar por OS para futuros relatórios
@app.route("/operador/amostragens")
def operador_listar():
    os_num = _norm(request.args.get("os"))
    part = _norm(request.args.get("partnumber"))
    op = _norm(request.args.get("operacao"))
    where = []
    params = []
    if os_num:
        where.append("a.os = %s")
        params.append(os_num)
    if part:
        where.append("a.partnumber = %s")
        params.append(part)
    if op:
        where.append("a.operacao = %s")
        params.append(op)

    sql = """
        SELECT a.id, a.os, a.partnumber, a.operacao, a.re_operador, a.created_at
        FROM operador_amostragem a
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY a.created_at DESC LIMIT 200"

    try:
        with _conn_db(DB_NAME) as c:
            with c.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": f"Falha ao consultar amostragens: {e}"}), 500

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "prep_path": PLANILHA_PREPARADOR_PATH,
        "oper_path": PLANILHA_OPERADOR_PATH
    })
def _mensagem_bloqueio(os_num: str, part: str, op: str, fonte: str, detalhe: str) -> str:
    """
    Gera um texto legível explicando por que o operador não pode registrar.
    fonte: "", "preparador_registro" ou "preparador_liberacao"
    detalhe: texto livre com dicas (ex.: "3/4 OK", "status_geral=pendente")
    """
    base = (
        "A máquina ainda não foi liberada pelo Preparador.\n"
        f"OS: {os_num}  •  Peça: {part}  •  Operação: {op}."
    )

    fonte = (fonte or "").strip().lower()
    det = str(detalhe or "")

    if not fonte:
        return base + "\nNão há registro do Preparador para esta combinação. Solicite a liberação (FOR-007/008)."

    if fonte == "preparador_liberacao":
        import re
        m = re.search(r"status_geral=([a-z_]+)", det, re.I)
        status = (m.group(1) if m else "pendente").replace("_", " ")
        return base + f"\nSituação da liberação: {status}. Procure o Preparador."

    if fonte == "preparador_registro":
        import re
        m = re.search(r"(\d+)\s*/\s*(\d+)", det)  # ex.: "3/4 OK"
        if m:
            ok_cnt, total = m.group(1), m.group(2)
            return base + f"\nProgresso do registro do Preparador: {ok_cnt}/{total} medidas OK. Aguarde até todas estarem OK."
        return base + "\nO registro do Preparador ainda não está 100% OK."

    return base

if __name__ == "__main__":
    # threaded=True mantém atendendo enquanto indexa em background entre requests
    app.run(host="0.0.0.0", port=5005, debug=True, threaded=True, use_reloader=False)
